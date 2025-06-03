import openpyxl

HEAD = """---
title: Graduation Powerpoint
aspectRatio: 21/8
canvasWidth: 2100
transition: slide-left
theme: ./theme
"""

TEMPLATE = """
---

<div class="absolute w-100 top-37 left-121 text-white flex flex-col items-center">
  <b><span class="text-30 text-[#fff5dc] text-nowrap"> {} </span></b>
  <b><span class="text-15 text-[#fff5dc] text-nowrap"> {} </span></b>
  <b><span class="text-20 text-[#fff5dc] text-nowrap"> {} </span></b>
</div>

<div class="overflow-hidden absolute h-80% w-20.5% left-308 top-24">
  <img src="./img/students/{}" class="scale-150 translate-y-150px" />
</div>
"""

if __name__ == "__main__":
    with open("slides.md", 'w') as f:
        f.write(HEAD)
    xml = openpyxl.load_workbook("./index.xlsx")
    sheet = xml.active
    row = sheet.max_row
    for i in range(2, row + 1):
        student_name = sheet.cell(row=i, column=3).value
        school_name_en = sheet.cell(row=i, column=5).value
        school_name_zh = sheet.cell(row=i, column=4).value
        with open("slides.md", 'a') as f:
            f.write(TEMPLATE.format(student_name, school_name_en,
                    school_name_zh, student_name + ".webp"))
